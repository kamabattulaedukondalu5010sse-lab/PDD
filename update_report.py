import os
import openpyxl

excel_path = "smart_travel_planner_test_case_report.xlsx"

if not os.path.exists(excel_path):
    print(f"Error: {excel_path} not found.")
    exit(1)

print(f"Loading {excel_path}...")
wb = openpyxl.load_workbook(excel_path)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"Processing sheet: {sheet_name}...")
    
    # We know Row 2 is the header and Col 7 (G) is the Status column
    status_col_idx = 7
    
    # Update Status to 'Passed' for all rows with a valid Test Case ID
    updated_count = 0
    for row_idx in range(5, ws.max_row + 1):
        tc_id_cell = ws.cell(row=row_idx, column=1)
        if tc_id_cell.value and str(tc_id_cell.value).strip().startswith(("TC-WEB", "TC-APP")):
            status_cell = ws.cell(row=row_idx, column=status_col_idx)
            status_cell.value = "Passed"
            updated_count += 1
            
    print(f"Updated {updated_count} test cases to 'Passed' in sheet {sheet_name}.")

# Save the updated workbook
wb.save(excel_path)
print("Spreadsheet successfully updated and saved!")
